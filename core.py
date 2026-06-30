"""
core.py
=======
Shared processing logic for the Placement Schedule Generator.
Used by both app.py (Streamlit) and process_placements.py (terminal).
"""

import io
import re
import zipfile
from datetime import date, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Excel helper ─────────────────────────────────────────────

def merge_range(row1, col1, row2, col2) -> str:
    return f"{get_column_letter(col1)}{row1}:{get_column_letter(col2)}{row2}"


# ── File loading ─────────────────────────────────────────────

def read_file(source) -> pd.DataFrame:
    """Read a CSV or Excel file from a path or file-like object."""
    name = getattr(source, "name", str(source))
    if str(name).lower().endswith(".csv"):
        return pd.read_csv(source)
    return pd.read_excel(source)


# ── InPlace loading and filtering ────────────────────────────

def load_inplace(source) -> pd.DataFrame:
    """
    Load the InPlace export. Expects these columns (position-independent):
      Student, Email, Agency, Start Date, Status, Placement Allocation Groups

    Note: Student Code is NOT required from InPlace — the Monash Student ID
    is sourced from the Google Form instead (see load_form / build_master).
    """
    df = read_file(source)

    required = ["Student", "Email", "Agency",
                "Start Date", "Status", "Placement Allocation Groups"]
    missing = [c for c in required if c not in df.columns]

    if missing:
        # Email is NOT included in iReport by default — it must be manually
        # ticked under iReport's "Add Column" option. Give a specific,
        # friendly explanation for this column rather than a generic
        # "missing columns" error, since this is the most common mistake
        # non-technical staff make.
        special_cols = [c for c in missing if c == "Email"]
        other_cols   = [c for c in missing if c != "Email"]

        message_parts = []

        if special_cols:
            message_parts.append(
                f"Your InPlace file is missing 'Email'.\n\n"
                f"This usually happens because Email is NOT included in "
                f"the iReport export by default — it must be added manually.\n\n"
                f"To fix this:\n"
                f"  1. Go back to iReport\n"
                f"  2. Find the 'Add Column' option (near the report filters or column settings)\n"
                f"  3. Tick the box next to 'Email'\n"
                f"  4. Apply the changes and download the report again\n\n"
                f"See the 'How to Download InPlace' tab in this app for a full walkthrough."
            )

        if other_cols:
            message_parts.append(
                f"Your InPlace file is also missing: {', '.join(other_cols)}.\n"
                f"Please check that you have downloaded the correct InPlace report "
                f"and that no columns were removed or renamed after downloading."
            )

        message_parts.append(
            f"\nColumns found in your file: {list(df.columns)}"
        )

        raise ValueError("\n\n".join(message_parts))

    df["Email"]         = df["Email"].astype(str).str.strip().str.lower()
    df["Agency"]        = df["Agency"].astype(str).str.strip()
    df["Student"]       = df["Student"].astype(str).str.strip()
    df["Start Date"]    = pd.to_datetime(df["Start Date"], errors="coerce")
    df["Status"]        = df["Status"].astype(str).str.strip()

    return df


def get_available_years(inplace_df: pd.DataFrame) -> list:
    """Extract unique years from the Start Date column, sorted descending."""
    years = (
        inplace_df["Start Date"]
        .dropna()
        .dt.year
        .unique()
    )
    return sorted([int(y) for y in years], reverse=True)


def _is_blank_text(series: pd.Series) -> pd.Series:
    """
    Robust blank check for a text column. Catches real NaN, empty string,
    and the literal text "nan"/"none"/"<na>" that astype(str) sometimes
    produces depending on column dtype and how the Excel/CSV file stored
    the empty cell.
    """
    s = series.astype(str).str.strip().str.lower()
    return s.isin(["", "nan", "none", "<na>"]) | s.isna()


def _status_rank(status_value, agency_value) -> int:
    """
    Priority rank for choosing a student's single 'winning' row when they
    have multiple rows. LOWER rank wins.

      0  Confirmed                  — beats everything
      1  Withdrawn (any variant)    — beats EDU and blanks
      2  Any other real status      — e.g. Offer
      3  EDU - anything             — low priority (no placement needed)
      4  Blank status               — lowest

    Note: EDU is detected from the Agency, since 'EDU - Credit from previous
    placement' etc. are agencies, not statuses.
    """
    s = str(status_value).strip().lower()
    a = str(agency_value).strip().lower()

    if s == "confirmed":
        return 0
    if s.startswith("withdrawn"):
        return 1
    if a.startswith("edu -"):
        return 3
    if s in ("", "nan", "none", "<na>"):
        return 4
    return 2


def _dedupe_by_student(df: pd.DataFrame) -> pd.DataFrame:
    """
    Collapse multiple rows per student (by Email) into one row, keeping
    the highest-priority row per _status_rank (Confirmed > Withdrawn >
    other > EDU > blank).
    """
    if len(df) == 0:
        return df
    df = df.copy()
    df["_rank"] = [
        _status_rank(s, a)
        for s, a in zip(df["Status"], df["Agency"])
    ]
    df = (
        df.sort_values("_rank", kind="stable")
          .drop_duplicates(subset="Email", keep="first")
          .drop(columns="_rank")
    )
    return df


def get_status_breakdown(inplace_df: pd.DataFrame, year: int) -> dict:
    """
    Count Confirmed / Withdrawn / No Placement / Other students for the
    selected year, BEFORE the Confirmed-only filter is applied. This gives
    staff visibility into the full picture, even though only Confirmed
    students (with a real Agency) are included in the final school exports.

    "No Placement" = student needs a placement but has not been assigned
    one yet: blank Status AND blank Agency (and NOT an 'EDU -' agency,
    since those students don't need a placement at all).

    Uses the same year-matching logic as filter_inplace (Requirement Groups
    or Placement Allocation Groups containing the year, with Start Date
    matching the year or being blank) so the counts are for the same
    student population that filter_inplace operates on.
    """
    df = inplace_df.copy()

    req_has_year = df["Requirement Groups"].astype(str).str.contains(
        str(year), na=False
    )
    alloc_has_year = df["Placement Allocation Groups"].astype(str).str.contains(
        str(year), na=False
    )
    group_year_ok = req_has_year | alloc_has_year

    start_year_ok = (
        df["Start Date"].dt.year == year
    ) | df["Start Date"].isna()

    df = df[group_year_ok & start_year_ok]

    # Deduplicate FIRST, across ALL rows (including EDU), so each student is
    # counted once by their winning row (Confirmed > Withdrawn > other >
    # EDU > blank). This matches how filter_inplace picks the winning row.
    df = _dedupe_by_student(df)

    status_lower = df["Status"].astype(str).str.lower()
    agency_lower = df["Agency"].astype(str).str.strip().str.lower()

    is_edu          = agency_lower.str.startswith("edu -", na=False)
    agency_blank    = _is_blank_text(df["Agency"])
    status_blank    = _is_blank_text(df["Status"])

    is_no_placement = agency_blank & status_blank & ~is_edu
    is_confirmed    = (status_lower == "confirmed")
    # Match "Withdrawn", "Withdrawn-Student", "Withdrawn-Agency", etc.
    is_withdrawn    = status_lower.str.startswith("withdrawn", na=False)

    # EDU students = those whose WINNING row is EDU (they have no Confirmed/
    # Withdrawn/other row that would have taken priority)
    n_edu          = int(is_edu.sum())
    n_no_placement = int((is_no_placement & ~is_edu).sum())
    n_confirmed    = int((is_confirmed & ~is_edu).sum())
    n_withdrawn    = int((is_withdrawn & ~is_edu).sum())
    n_other        = len(df) - n_confirmed - n_withdrawn - n_no_placement - n_edu

    return {
        "confirmed":     n_confirmed,
        "withdrawn":     n_withdrawn,
        "no_placement":  n_no_placement,
        "other":         int(n_other),
        "edu":           n_edu,
    }


def get_no_placement_students(inplace_df: pd.DataFrame, year: int) -> pd.DataFrame:
    """
    Return InPlace rows for students who need a placement but have not
    been assigned one yet (blank Status AND blank Agency, excluding
    'EDU -' agencies which mean no placement is required at all).

    Uses the same year-matching logic as filter_inplace.
    """
    df = inplace_df.copy()

    req_has_year = df["Requirement Groups"].astype(str).str.contains(
        str(year), na=False
    )
    alloc_has_year = df["Placement Allocation Groups"].astype(str).str.contains(
        str(year), na=False
    )
    group_year_ok = req_has_year | alloc_has_year

    start_year_ok = (
        df["Start Date"].dt.year == year
    ) | df["Start Date"].isna()

    df = df[group_year_ok & start_year_ok]

    # Dedupe FIRST so each student is represented by their winning row.
    # A student whose winning row is Confirmed/Withdrawn/EDU will NOT be
    # treated as no-placement; only those whose winning row is blank+blank.
    df = _dedupe_by_student(df)

    df = df[~df["Agency"].str.startswith("EDU -", na=False)]

    agency_blank = _is_blank_text(df["Agency"])
    status_blank = _is_blank_text(df["Status"])

    return df[agency_blank & status_blank].reset_index(drop=True)


def filter_inplace(inplace_df: pd.DataFrame, year: int):
    """
    Apply all filters to the InPlace data:

      Year filter (two-pronged — handles unplaced students with blank Start Date):
        Keep row if EITHER:
          a) Start Date year == selected year, OR
          b) Start Date is blank/null
        AND Requirement Groups contains the selected year as a 4-digit number.
        This ensures unplaced students (no Start Date yet) are included as long
        as their Requirement Group is for the correct year.

      Then:
        Exclude agencies starting with 'EDU -'
        Keep only Status == 'Confirmed' (Withdrawn, Offer, blank, etc. excluded)

    Returns:
      filtered_df   — the filtered InPlace DataFrame
      window_start  — earliest Start Date in this year (date), or None
      window_end    — latest End Date in this year (date), or None
    """
    import re as _re
    df = inplace_df.copy()

    # Year in Requirement Groups (catches all rows including unplaced)
    req_has_year = df["Requirement Groups"].astype(str).str.contains(
        str(year), na=False
    )
    # Also accept Placement Allocation Groups as a fallback
    alloc_has_year = df["Placement Allocation Groups"].astype(str).str.contains(
        str(year), na=False
    )
    group_year_ok = req_has_year | alloc_has_year

    # Start Date: matches year OR is blank
    start_year_ok = (
        df["Start Date"].dt.year == year
    ) | df["Start Date"].isna()

    # Both conditions must be true
    df = df[group_year_ok & start_year_ok]

    # Capture the placement window from rows that DO have a Start Date
    placed = df[df["Start Date"].notna()]
    window_start = placed["Start Date"].min().date() if len(placed) > 0 else None
    end_col = pd.to_datetime(placed["End Date"], errors="coerce") if len(placed) > 0 else pd.Series([], dtype="datetime64[ns]")
    window_end = end_col.max().date() if len(placed) > 0 and pd.notna(end_col.max()) else None

    # Deduplicate FIRST, across ALL rows (including EDU rows), so that a
    # student with e.g. an 'EDU - Credit' row AND a Confirmed row keeps the
    # Confirmed row. Confirmed beats everything; Withdrawn beats EDU; an
    # EDU-only student keeps their EDU row (and is then excluded below).
    df = _dedupe_by_student(df)

    # Now exclude EDU - agencies (students who don't need a placement —
    # this only removes students whose WINNING row is EDU, i.e. they have
    # no Confirmed/Withdrawn/other row)
    df = df[~df["Agency"].str.startswith("EDU -", na=False)]

    # Keep only Confirmed status (Withdrawn, Offer, blank, etc. are excluded)
    df = df[df["Status"].str.lower() == "confirmed"]

    return df.reset_index(drop=True), window_start, window_end


def clean_school_name(name) -> str:
    """Normalise school name — same logic as before."""
    if not isinstance(name, str) or name.strip() == "":
        return name
    s = name.strip()
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'\s*-\s*', ' - ', s)
    s = s.title()
    return s


# ── Google Form loading ──────────────────────────────────────

def load_form(source) -> pd.DataFrame:
    """
    Load the Google Form CSV/Excel export.

    Matches columns by NAME (robust to reordering) rather than by position.
    The current form has these columns:
      Timestamp, Email address, Full Name,
      School Name - as written on InPlace:,
      What days will you be attending placement?,
      I confirm that I have double checked my days...

    Only Timestamp, Email, Full Name, and the days column are used.
    """
    df = read_file(source)

    def find_col(*keywords):
        """Find the first column whose lowercased name contains all keywords."""
        for col in df.columns:
            name = str(col).lower()
            if all(k in name for k in keywords):
                return col
        return None

    ts_col    = find_col("timestamp")
    email_col = find_col("email")
    name_col  = find_col("full name") or find_col("name")
    days_col  = find_col("days") or find_col("attending")

    missing = []
    if ts_col is None:    missing.append("Timestamp")
    if email_col is None: missing.append("Email address")
    if name_col is None:  missing.append("Full Name")
    if days_col is None:  missing.append("What days will you be attending placement?")
    if missing:
        raise ValueError(
            f"The Google Form file is missing these column(s): {', '.join(missing)}.\n"
            f"Columns found in your file: {list(df.columns)}"
        )

    out = pd.DataFrame()
    out["timestamp"] = df[ts_col]
    out["email"]     = df[email_col].astype(str).str.strip().str.lower()
    out["full_name"] = df[name_col].astype(str).str.strip()
    out["days_raw"]  = df[days_col].astype(str).str.strip()
    out["days_raw"]  = out["days_raw"].where(out["days_raw"] != "nan", other="")

    return out


# ── Date detection ───────────────────────────────────────────

def extract_year_from_form(form_df: pd.DataFrame) -> int:
    years = []
    for ts in form_df["timestamp"].dropna():
        try:
            years.append(int(str(ts).split("/")[2].split(" ")[0]))
        except (IndexError, ValueError):
            pass
    if not years:
        raise ValueError("Could not detect year from the Form timestamp column.")
    return max(set(years), key=years.count)


def parse_form_day(raw_string: str, year: int):
    s = raw_string.strip()
    s = re.sub(r"(\d+)(st|nd|rd|th)\b", r"\1", s)
    s = re.sub(
        r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s+",
        "", s
    ).strip()
    s = f"{s} {year}"
    for fmt in ("%B %d %Y", "%d %B %Y"):
        try:
            p = pd.to_datetime(s, format=fmt)
            return date(p.year, p.month, p.day)
        except Exception:
            pass
    return None


def detect_placement_dates(form_df: pd.DataFrame, year: int):
    """
    Parse placement dates from the form using the timestamp year as the
    placement year. Students always submit one month before a mid-year
    placement, so the timestamp year always matches the placement year.
    This also acts as a natural per-year filter — only tokens that parse
    into a weekday in the correct year are included.
    """
    # Filter form rows to only those submitted in the selected year
    year_mask = pd.to_datetime(form_df["timestamp"], dayfirst=True, errors="coerce").dt.year == year
    form_year = form_df[year_mask]

    all_tokens = set()
    for cell in form_year["days_raw"]:
        if cell:
            for token in cell.split(","):
                all_tokens.add(token.strip())

    sampled = set()
    for token in all_tokens:
        d = parse_form_day(token, year)
        if d and d.weekday() < 5:
            sampled.add(d)

    if not sampled:
        raise ValueError(
            f"No valid placement dates found in the Form for {year}. "
            f"Check that the form responses include submissions from {year}."
        )

    ordered_dates = []
    cursor = min(sampled)
    while len(ordered_dates) < 15:
        if cursor.weekday() < 5:
            ordered_dates.append(cursor)
        cursor += timedelta(days=1)

    day_lookup = {}
    for token in all_tokens:
        d = parse_form_day(token, year)
        if d and d in ordered_dates:
            day_lookup[token] = d

    return ordered_dates, day_lookup


def build_schedule_labels(ordered_dates: list):
    short_labels = [
        f"{d.strftime('%a')} {d.day} {d.strftime('%b')}"
        for d in ordered_dates
    ]
    week_of = [i // 5 + 1 for i in range(15)]
    week_labels = []
    for w in range(1, 4):
        idx = [i for i, x in enumerate(week_of) if x == w]
        d1, d2 = ordered_dates[idx[0]], ordered_dates[idx[-1]]
        week_labels.append(
            f"Week {w}  ({d1.strftime('%b')} {d1.day}"
            f" - {d2.strftime('%b')} {d2.day})"
        )
    range_label = (
        f"{ordered_dates[0].strftime('%b')} {ordered_dates[0].day} - "
        f"{ordered_dates[14].strftime('%b')} {ordered_dates[14].day} "
        f"{ordered_dates[14].year}"
    )
    return short_labels, week_of, week_labels, range_label


# ── Master join ──────────────────────────────────────────────

def _build_form_lookup(form_df: pd.DataFrame, day_lookup: dict) -> dict:
    """
    Build a lookup: email -> {days set, submitted}

    Since form_df is already filtered by timestamp year in
    detect_placement_dates, only the current year's submissions are
    considered here. If a student submitted multiple times in the same
    year, the submission with the most days in the placement window wins.
    """
    form_lookup = {}
    for _, row in form_df.iterrows():
        email = row["email"]
        days = set()
        if row["days_raw"]:
            for token in row["days_raw"].split(","):
                token = token.strip()
                if token in day_lookup:
                    days.add(day_lookup[token])

        if days:
            existing = form_lookup.get(email, {})
            if len(days) >= len(existing.get("days", set())):
                form_lookup[email] = {
                    "days":      days,
                    "submitted": True,
                }
    return form_lookup


def build_master(inplace_df: pd.DataFrame,
                 form_df: pd.DataFrame,
                 day_lookup: dict) -> pd.DataFrame:
    """
    Join InPlace (already filtered to Confirmed students with a real
    Agency) with Form responses on email.

    Returns:
        master_df — one row per student, with all fields + date columns
    """
    form_lookup = _build_form_lookup(form_df, day_lookup)

    master_rows = []
    for _, ip_row in inplace_df.iterrows():
        email      = ip_row["Email"].lower().strip()
        form_data  = form_lookup.get(email, {"days": set(), "submitted": False})
        master_rows.append({
            "student_name": ip_row["Student"],
            "email":        email,
            "agency":       clean_school_name(ip_row["Agency"]),
            "submitted":    form_data["submitted"],
            "days":         form_data["days"],
        })

    return pd.DataFrame(master_rows)


def build_no_placement_df(no_placement_inplace: pd.DataFrame,
                           form_df: pd.DataFrame,
                           day_lookup: dict) -> pd.DataFrame:
    """
    Build the "No Placement" student list directly from InPlace rows
    (blank Status AND blank Agency, excluding 'EDU -' — see
    get_no_placement_students). Each student is enriched with their Form
    submission (selected days) if they happen to have submitted the form
    already, even though they have no placement yet.
    """
    form_lookup = _build_form_lookup(form_df, day_lookup)

    rows = []
    for _, ip_row in no_placement_inplace.iterrows():
        email     = ip_row["Email"].lower().strip()
        form_data = form_lookup.get(email, {"days": set(), "submitted": False})
        rows.append({
            "student_name": ip_row["Student"],
            "email":        email,
            "submitted":    form_data["submitted"],
            "days":         form_data["days"],
        })

    return pd.DataFrame(rows)


# ── Build attendance matrix for one school ───────────────────

def build_matrix_from_master(school: str,
                              master_df: pd.DataFrame,
                              ordered_dates: list) -> pd.DataFrame:
    """
    Returns a DataFrame for one school:
    - Submitted students first (sorted by name), with tick marks
    - Non-submitted students at the bottom, flagged
    """
    school_df = master_df[master_df["agency"] == school].copy()

    submitted     = school_df[school_df["submitted"]].sort_values("student_name")
    not_submitted = school_df[~school_df["submitted"]].sort_values("student_name")

    rows = []

    for _, s in submitted.iterrows():
        row = {
            "full_name":    s["student_name"],
            "email":        s["email"],
            "note":         "",
        }
        for d in ordered_dates:
            row[d] = "✓" if d in s["days"] else ""
        rows.append(row)

    for _, s in not_submitted.iterrows():
        row = {
            "full_name":    s["student_name"],
            "email":        s["email"],
            "note":         "No preference submitted",
        }
        for d in ordered_dates:
            row[d] = ""
        rows.append(row)

    return pd.DataFrame(rows).reset_index(drop=True)


# ── Colour palette ───────────────────────────────────────────

NAVY      = "0D47A1"
DARK_NAVY = "1F3864"
MID_NAVY  = "1565C0"
DARK_GREY = "2C3E50"
TICK_BG   = "00897B"
ALT_ROW   = "F5F5F5"
NO_SUB_BG = "FFF3CD"   # amber tint for "no preference submitted" rows
NO_SUB_FG = "856404"

WK_BG  = ["FFF9C4", "FCE4EC", "E8EAF6"]
WK_HDR = ["F9A825", "AD1457", "5C6BC0"]

COUNT_GREEN = ("2ECC71", "27AE60")
COUNT_AMBER = ("F39C12", "E67E22")
COUNT_RED   = ("E74C3C", "C0392B")


def make_border(colour="CCCCCC"):
    s = Side(style="thin", color=colour)
    return Border(left=s, right=s, top=s, bottom=s)

def solid_fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def style_cell(cell, bg=None, fg="000000", bold=False, size=10,
               h_align="center", border_colour="CCCCCC",
               wrap=False, italic=False):
    if bg:
        cell.fill = solid_fill(bg)
    cell.font      = Font(name="Arial", bold=bold, italic=italic,
                          color=fg, size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical="center",
                               wrap_text=wrap)
    cell.border    = make_border(border_colour)


# ── Build one Excel workbook ─────────────────────────────────

def build_workbook(school, mat, ordered_dates, short_labels,
                   week_of, week_labels, range_label) -> Workbook:

    n_students    = len(mat)
    # Columns: Full Name | Email | Note | [days]
    fixed_headers = ["Full Name", "Email Address", "Note"]
    n_fixed       = len(fixed_headers)   # 3
    n_days        = 15
    total_cols    = n_fixed + n_days

    wb = Workbook()

    # ── Sheet 1: Attendance Schedule ────────────────────────
    ws = wb.active
    ws.title = "Attendance Schedule"
    ws.sheet_view.showGridLines = False

    # Row 1: school title
    ws.merge_cells(merge_range(1, 1, 1, total_cols))
    c = ws.cell(1, 1, school)
    style_cell(c, bg=NAVY, fg="FFFFFF", bold=True, size=14,
               border_colour="FFFFFF")
    ws.row_dimensions[1].height = 32

    # Row 2: subtitle
    ws.merge_cells(merge_range(2, 1, 2, total_cols))
    c = ws.cell(2, 1,
        f"Placement Attendance  |  {n_students} student(s)  |  "
        f"3-Week Period: {range_label}  |  10 days per student")
    style_cell(c, bg=MID_NAVY, fg="E3F2FD", size=11,
               wrap=True, border_colour="FFFFFF")
    ws.row_dimensions[2].height = 22

    # Row 3: week group headers
    for col in range(1, n_fixed + 1):
        style_cell(ws.cell(3, col), bg=DARK_NAVY, border_colour="FFFFFF")
    for w in range(3):
        cols = [i for i, x in enumerate(week_of) if x == w + 1]
        cs = n_fixed + cols[0] + 1
        ce = n_fixed + cols[-1] + 1
        ws.merge_cells(merge_range(3, cs, 3, ce))
        c = ws.cell(3, cs, week_labels[w])
        style_cell(c, bg=WK_HDR[w], fg="FFFFFF", bold=True,
                   border_colour="FFFFFF")
    ws.row_dimensions[3].height = 22

    # Row 4: column headers
    for ci, h in enumerate(fixed_headers + short_labels, 1):
        c = ws.cell(4, ci, h)
        style_cell(c, bg=DARK_NAVY, fg="FFFFFF", bold=True,
                   size=11, border_colour="FFFFFF", wrap=True)
    ws.row_dimensions[4].height = 36

    # Rows 5+: student data
    data_start = 5
    for r_idx, row in mat.iterrows():
        rn          = data_start + r_idx
        no_sub      = row["note"] == "No preference submitted"
        row_bg      = NO_SUB_BG if no_sub else (ALT_ROW if r_idx % 2 == 1 else None)
        text_colour = NO_SUB_FG if no_sub else "000000"

        for ci, val in enumerate(
            [row["full_name"], row["email"], row["note"]], 1
        ):
            c = ws.cell(rn, ci, val)
            style_cell(c, bg=row_bg, fg=text_colour,
                        h_align="left" if ci <= 2 else "center",
                        italic=no_sub and ci == 3)

        for di, d in enumerate(ordered_dates):
            ci  = n_fixed + di + 1
            val = row[d]
            c   = ws.cell(rn, ci, val)
            if val == "✓":
                style_cell(c, bg=TICK_BG, fg="FFFFFF", bold=True, size=11)
            else:
                style_cell(c, bg=NO_SUB_BG if no_sub else WK_BG[week_of[di] - 1])
        ws.row_dimensions[rn].height = 18

    # Total row (only counts submitted students)
    tr = data_start + n_students
    ws.merge_cells(merge_range(tr, 1, tr, n_fixed))
    c = ws.cell(tr, 1, "TOTAL STUDENTS PER DAY")
    style_cell(c, bg=DARK_GREY, fg="FFFFFF", bold=True,
               h_align="left", border_colour="FFFFFF")
    for di, d in enumerate(ordered_dates):
        n_att = (mat[d] == "✓").sum()
        c = ws.cell(tr, n_fixed + di + 1, n_att)
        style_cell(c, bg=DARK_NAVY, fg="FFFFFF", bold=True,
                   border_colour="FFFFFF")
    ws.row_dimensions[tr].height = 22

    # Column widths and freeze
    ws.column_dimensions["A"].width = 26   # full name
    ws.column_dimensions["B"].width = 32   # email
    ws.column_dimensions["C"].width = 26   # note
    for di in range(n_days):
        ws.column_dimensions[get_column_letter(n_fixed + di + 1)].width = 10
    ws.freeze_panes = ws.cell(data_start, n_fixed + 1)

    # ── Sheet 2: Daily Summary ───────────────────────────────
    ws2 = wb.create_sheet("Daily Summary")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D1")
    c = ws2.cell(1, 1, f"{school} -- Daily Student Count")
    style_cell(c, bg=NAVY, fg="FFFFFF", bold=True, size=13,
               border_colour="FFFFFF")
    ws2.row_dimensions[1].height = 28

    for ci, h in enumerate(
        ["Week", "Date", "Day of Week", "Students Attending"], 1
    ):
        c = ws2.cell(2, ci, h)
        style_cell(c, bg=DARK_NAVY, fg="FFFFFF", bold=True,
                   size=11, border_colour="FFFFFF")
    ws2.row_dimensions[2].height = 22

    for di, d in enumerate(ordered_dates):
        r   = 3 + di
        w   = week_of[di] - 1
        n   = (mat[d] == "✓").sum()
        ws2.cell(r, 1, week_labels[w])
        ws2.cell(r, 2, short_labels[di])
        ws2.cell(r, 3, d.strftime("%A"))
        ws2.cell(r, 4, n)
        for ci in range(1, 4):
            style_cell(ws2.cell(r, ci), bg=WK_BG[w], h_align="left")
        bg_c, bdr_c = (COUNT_GREEN if n >= 5
                       else COUNT_AMBER if n >= 2
                       else COUNT_RED)
        style_cell(ws2.cell(r, 4), bg=bg_c, fg="FFFFFF",
                   bold=True, border_colour=bdr_c)
        ws2.row_dimensions[r].height = 20

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 22

    note_row = 3 + n_days + 1
    ws2.merge_cells(f"A{note_row}:D{note_row}")
    c = ws2.cell(note_row, 1,
                 "Colour key:  Green = 5+ students   "
                 "Amber = 2-4 students   Red = 0-1 students")
    c.font      = Font(name="Arial", size=9, color="555555", italic=True)
    c.alignment = Alignment(horizontal="left")

    return wb


# ── Build "No Placement" workbook ─────────────────────────────

def build_no_placement_workbook(no_placement_df: pd.DataFrame,
                                 ordered_dates: list,
                                 short_labels: list,
                                 week_of: list,
                                 week_labels: list,
                                 range_label: str) -> Workbook:
    """
    Workbook listing students who need a placement but have not been
    assigned one yet in InPlace (blank Status and blank Agency).
    If a student has also submitted the Google Form, their preferred
    name and selected days are shown too.
    """

    n_fixed    = 3   # Full Name | Email | Submitted Form?
    n_days     = 15
    total_cols = n_fixed + n_days

    wb = Workbook()
    ws = wb.active
    ws.title = "No Placement"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(merge_range(1, 1, 1, total_cols))
    c = ws.cell(1, 1, "Students Awaiting Placement Assignment")
    style_cell(c, bg=NAVY, fg="FFFFFF", bold=True, size=14,
               border_colour="FFFFFF")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(merge_range(2, 1, 2, total_cols))
    c = ws.cell(2, 1,
        f"These students are in InPlace with no Status and no Agency assigned yet "
        f"(placement not yet confirmed)  |  Period: {range_label}")
    style_cell(c, bg=MID_NAVY, fg="E3F2FD", size=11,
               wrap=True, border_colour="FFFFFF")
    ws.row_dimensions[2].height = 22

    # Week headers
    for col in range(1, n_fixed + 1):
        style_cell(ws.cell(3, col), bg=DARK_NAVY, border_colour="FFFFFF")
    for w in range(3):
        cols = [i for i, x in enumerate(week_of) if x == w + 1]
        cs = n_fixed + cols[0] + 1
        ce = n_fixed + cols[-1] + 1
        ws.merge_cells(merge_range(3, cs, 3, ce))
        c = ws.cell(3, cs, week_labels[w])
        style_cell(c, bg=WK_HDR[w], fg="FFFFFF", bold=True,
                   border_colour="FFFFFF")
    ws.row_dimensions[3].height = 22

    # Column headers
    headers = ["Full Name", "Email Address", "Submitted Form?"] + short_labels
    for ci, h in enumerate(headers, 1):
        c = ws.cell(4, ci, h)
        style_cell(c, bg=DARK_NAVY, fg="FFFFFF", bold=True,
                   size=11, border_colour="FFFFFF", wrap=True)
    ws.row_dimensions[4].height = 36

    data_start = 5
    for r_idx, row in no_placement_df.reset_index(drop=True).iterrows():
        rn         = data_start + r_idx
        submitted  = row.get("submitted", False)
        row_bg     = ALT_ROW if r_idx % 2 == 1 else None

        fixed_vals = [
            row.get("student_name", ""), row.get("email", ""),
            "Yes" if submitted else "No",
        ]
        for ci, val in enumerate(fixed_vals, 1):
            c = ws.cell(rn, ci, val)
            if ci == 3:
                style_cell(c, bg=row_bg,
                           fg="1B5E20" if submitted else "999999",
                           bold=submitted, h_align="center")
            else:
                style_cell(c, bg=row_bg, h_align="left")

        for di, d in enumerate(ordered_dates):
            ci  = n_fixed + di + 1
            val = "✓" if d in row.get("days", set()) else ""
            c   = ws.cell(rn, ci, val)
            if val == "✓":
                style_cell(c, bg=TICK_BG, fg="FFFFFF", bold=True, size=11)
            else:
                style_cell(c, bg=WK_BG[week_of[di] - 1])
        ws.row_dimensions[rn].height = 18

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    for di in range(n_days):
        ws.column_dimensions[get_column_letter(n_fixed + di + 1)].width = 10
    ws.freeze_panes = ws.cell(data_start, n_fixed + 1)

    return wb


# ── Save to bytes ────────────────────────────────────────────

def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_zip(schools_data: list) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, wb_bytes in schools_data:
            zf.writestr(filename, wb_bytes)
    return buf.getvalue()
