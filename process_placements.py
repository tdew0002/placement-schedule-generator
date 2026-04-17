"""
process_placements.py
=====================
Reads a Google Form CSV export of student placement availability
and produces one formatted Excel workbook per school.

HOW TO USE (every year):
  1. Put this file and Source.csv in the same folder
  2. Open terminal in this folder (with venv active) and run:
       python3 process_placements.py
  3. Choose which schools to export when prompted
  4. Find your Excel files in the school_schedules/ folder

No dates to enter — everything is read automatically from the CSV.

REQUIREMENTS (install once):
  pip install pandas openpyxl
"""

import os
import re
from datetime import date, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Helper: row/col numbers -> Excel range string e.g. "A1:S1"
# Needed because newer openpyxl only accepts string ranges.
# ============================================================


def merge_range(row1, col1, row2, col2) -> str:
    return f"{get_column_letter(col1)}{row1}:{get_column_letter(col2)}{row2}"


# ============================================================
# Step 1: Read the CSV and detect everything automatically
# ============================================================


def clean_school(name) -> str:
    """
    Normalise a school name so that variations that look the same
    on screen are treated as identical. Handles the three most
    common issues found in Google Form exports:
      - Leading / trailing spaces  e.g. ' Mazenod College'
      - Inconsistent capitalisation  e.g. 'Mazenod college'
      - Irregular spacing around dashes  e.g. 'Mazenod College- Secondary'
    """
    if not isinstance(name, str):
        return name
    s = name.strip()
    s = re.sub(r"\s+", " ", s)  # collapse multiple spaces to one
    s = re.sub(r"\s*-\s*", " - ", s)  # normalise spaces around dashes
    s = s.title()  # Title Case every word
    return s


def load_csv(input_file: str) -> pd.DataFrame:
    raw = pd.read_csv(input_file)
    raw.columns = [
        "timestamp",
        "email",
        "full_name",
        "pref_name",
        "student_id",
        "school",
        "days_raw",
        *raw.columns[7:],
    ]
    raw["display_name"] = raw.apply(
        lambda r: (
            str(r["pref_name"]).strip()
            if pd.notna(r["pref_name"])
            and str(r["pref_name"]).strip() not in ("", "nan")
            else str(r["full_name"]).strip()
        ),
        axis=1,
    )
    raw["full_name"] = raw["full_name"].str.strip()
    raw["email"] = raw["email"].str.strip()
    raw["school"] = raw["school"].apply(clean_school)
    return raw


def extract_year_from_timestamps(raw: pd.DataFrame) -> int:
    """
    Pull the placement year from the Google Form timestamp column.
    The format Google Forms uses is DD/MM/YYYY HH:MM:SS.
    We take the most common year found across all rows.
    """
    years = []
    for ts in raw["timestamp"].dropna():
        try:
            # "18/06/2025 08:04:00" -> split on / -> [18, 06, 2025 08:04:00]
            year_part = str(ts).split("/")[2].split(" ")[0]
            years.append(int(year_part))
        except (IndexError, ValueError):
            pass

    if not years:
        raise SystemExit(
            "ERROR: Could not read dates from the Timestamp column.\n"
            "Check that the CSV is a Google Form export and has not been edited."
        )

    # Use the most common year (handles edge cases where a handful
    # of timestamps might be from a different year)
    return max(set(years), key=years.count)


def parse_form_day(raw_string: str, year: int):
    """
    Parse a Google Form day string into a date object.
    Handles formats like:
      "Monday July 21"   "Friday 1st August"   "Thursday July 31st"
    """
    s = raw_string.strip()
    s = re.sub(r"(\d+)(st|nd|rd|th)\b", r"\1", s)  # strip ordinals
    s = re.sub(  # strip weekday name
        r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s+", "", s
    ).strip()
    s = f"{s} {year}"
    for fmt in ("%B %d %Y", "%d %B %Y"):
        try:
            p = pd.to_datetime(s, format=fmt)
            return date(p.year, p.month, p.day)
        except Exception:
            pass
    return None


def detect_placement_dates(raw: pd.DataFrame, year: int):
    """
    Parse every day token from the form responses, keep only weekdays,
    then use the earliest date to generate the full 15-day Mon-Fri grid.

    Why generate from the earliest rather than just using what's present:
    Students pick 10 of 15 days, so some days may have zero selections
    and won't appear in the CSV at all. Generating forward from the
    earliest date guarantees all 15 days are always included.
    """
    all_tokens = set()
    for cell in raw["days_raw"].dropna():
        for token in str(cell).split(","):
            all_tokens.add(token.strip())

    # Parse every token to a date, keep only Mon-Fri
    sampled = set()
    day_lookup = {}
    for token in all_tokens:
        d = parse_form_day(token, year)
        if d and d.weekday() < 5:
            sampled.add(d)
            day_lookup[token] = d  # temporary; rebuilt after grid is finalised

    if not sampled:
        raise SystemExit(
            "ERROR: No valid placement dates found in the CSV.\n"
            "Check that the 'What days will you be attending' column is present."
        )

    # Generate 15 consecutive weekdays starting from the earliest date
    ordered_dates = []
    cursor = min(sampled)
    while len(ordered_dates) < 15:
        if cursor.weekday() < 5:
            ordered_dates.append(cursor)
        cursor += timedelta(days=1)

    # Rebuild day_lookup keeping only tokens that map to a grid date
    day_lookup = {}
    for token in all_tokens:
        d = parse_form_day(token, year)
        if d and d in ordered_dates:
            day_lookup[token] = d

    return ordered_dates, day_lookup


# ============================================================
# Step 2: School selection prompt
# ============================================================


def ask_school_selection(schools: list) -> list:
    print("School selection")
    print("-" * 40)
    print(f"  {len(schools)} schools found in the CSV.")
    print()
    print("  Options:")
    print("    A  -  Export ALL schools")
    print("    S  -  Search and select specific schools")
    print()
    while True:
        choice = input("  Your choice (A / S): ").strip().upper()
        if choice == "A":
            print(f"\n  Exporting all {len(schools)} schools.\n")
            return schools
        elif choice == "S":
            return ask_pick_schools(schools)
        else:
            print("  Please type A or S.")


def ask_pick_schools(schools: list) -> list:
    selected = []
    print()
    print("  Type part of a school name to search.")
    print("  You can add multiple schools one search at a time.")
    print("  Type DONE when finished, or ALL to switch to all schools.")
    print()

    while True:
        query = input("  Search (or DONE / ALL): ").strip()

        if query.upper() == "ALL":
            print(f"\n  Switched to all {len(schools)} schools.\n")
            return schools

        if query.upper() == "DONE":
            if not selected:
                print("  No schools selected yet. Search for at least one first.")
                continue
            print(f"\n  {len(selected)} school(s) selected:\n")
            for s in selected:
                print(f"    - {s}")
            print()
            return selected

        if not query:
            continue

        matches = [s for s in schools if query.lower() in s.lower()]

        if not matches:
            print(f"  No schools found matching '{query}'. Try again.\n")
            continue

        if len(matches) == 1:
            school = matches[0]
            if school in selected:
                print(f"  '{school}' is already in your selection.\n")
            else:
                selected.append(school)
                print(f"  Added: {school}")
                print(f"  Selection so far: {len(selected)} school(s)\n")
            continue

        print(f"\n  {len(matches)} schools found:\n")
        for i, s in enumerate(matches, 1):
            tag = " (already selected)" if s in selected else ""
            print(f"    {i:>3}.  {s}{tag}")
        print()
        print("  Enter a number to add that school, or press Enter to search again.")
        print()

        while True:
            pick = input("  Number (or Enter to skip): ").strip()
            if pick == "":
                break
            try:
                idx = int(pick) - 1
                if not 0 <= idx < len(matches):
                    raise ValueError
                school = matches[idx]
                if school in selected:
                    print(f"  '{school}' is already selected.\n")
                else:
                    selected.append(school)
                    print(f"  Added: {school}")
                    print(f"  Selection so far: {len(selected)} school(s)\n")
                break
            except (ValueError, IndexError):
                print(f"  Please enter a number between 1 and {len(matches)}.")
        print()


# ============================================================
# Excel styling helpers
# ============================================================


def make_border(colour="CCCCCC"):
    s = Side(style="thin", color=colour)
    return Border(left=s, right=s, top=s, bottom=s)


def solid_fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)


def style_cell(
    cell,
    bg=None,
    fg="000000",
    bold=False,
    size=10,
    h_align="center",
    border_colour="CCCCCC",
    wrap=False,
    italic=False,
):
    if bg:
        cell.fill = solid_fill(bg)
    cell.font = Font(name="Arial", bold=bold, italic=italic, color=fg, size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)
    cell.border = make_border(border_colour)


# ============================================================
# Colour palette
# ============================================================

NAVY = "0D47A1"
DARK_NAVY = "1F3864"
MID_NAVY = "1565C0"
DARK_GREY = "2C3E50"
TICK_BG = "00897B"
ALT_ROW = "F5F5F5"

WK_BG = ["FFF9C4", "FCE4EC", "E8EAF6"]  # yellow, pink, lavender
WK_HDR = ["F9A825", "AD1457", "5C6BC0"]  # amber, deep pink, indigo

COUNT_GREEN = ("2ECC71", "27AE60")
COUNT_AMBER = ("F39C12", "E67E22")
COUNT_RED = ("E74C3C", "C0392B")


# ============================================================
# Build one Excel workbook for a single school
# ============================================================


def build_workbook(
    school, mat, ordered_dates, short_labels, week_of, week_labels, range_label
):

    n_students = len(mat)
    fixed_headers = ["Preferred Name", "Full Name", "Email Address", "Student ID"]
    n_fixed = 4
    n_days = 15
    total_cols = n_fixed + n_days

    wb = Workbook()

    # ---- Sheet 1: Attendance Schedule ----------------------
    ws = wb.active
    ws.title = "Attendance Schedule"
    ws.sheet_view.showGridLines = False

    # Row 1: school title
    ws.merge_cells(merge_range(1, 1, 1, total_cols))
    c = ws.cell(1, 1, school)
    style_cell(c, bg=NAVY, fg="FFFFFF", bold=True, size=14, border_colour="FFFFFF")
    ws.row_dimensions[1].height = 32

    # Row 2: subtitle
    ws.merge_cells(merge_range(2, 1, 2, total_cols))
    c = ws.cell(
        2,
        1,
        f"Placement Attendance  |  {n_students} student(s)  |  "
        f"3-Week Period: {range_label}  |  10 days per student",
    )
    style_cell(c, bg=MID_NAVY, fg="E3F2FD", size=11, wrap=True, border_colour="FFFFFF")
    ws.row_dimensions[2].height = 22

    # Row 3: week group headers
    for col in range(1, n_fixed + 1):
        style_cell(ws.cell(3, col), bg=DARK_NAVY, border_colour="FFFFFF")
    for w in range(3):
        cols = [i for i, x in enumerate(week_of) if x == w + 1]
        cs = n_fixed + cols[0] + 1
        ce = n_fixed + cols[-1] + 1
        ws.merge_cells(merge_range(3, cs, 3, ce))
        c = ws.cell(3, cs, week_labels[w])
        style_cell(c, bg=WK_HDR[w], fg="FFFFFF", bold=True, border_colour="FFFFFF")
    ws.row_dimensions[3].height = 22

    # Row 4: column headers
    for ci, h in enumerate(fixed_headers + short_labels, 1):
        c = ws.cell(4, ci, h)
        style_cell(
            c,
            bg=DARK_NAVY,
            fg="FFFFFF",
            bold=True,
            size=11,
            border_colour="FFFFFF",
            wrap=True,
        )
    ws.row_dimensions[4].height = 36

    # Rows 5+: student data
    data_start = 5
    for r_idx, row in mat.iterrows():
        rn = data_start + r_idx
        row_bg = ALT_ROW if r_idx % 2 == 1 else None

        for ci, val in enumerate(
            [row["display_name"], row["full_name"], row["email"], row["student_id"]], 1
        ):
            c = ws.cell(rn, ci, val)
            style_cell(c, bg=row_bg, h_align="left" if ci <= 3 else "center")

        for di, d in enumerate(ordered_dates):
            ci = n_fixed + di + 1
            val = row[d]
            c = ws.cell(rn, ci, val)
            if val == "✓":
                style_cell(c, bg=TICK_BG, fg="FFFFFF", bold=True, size=11)
            else:
                style_cell(c, bg=WK_BG[week_of[di] - 1])
        ws.row_dimensions[rn].height = 18

    # Total row
    tr = data_start + n_students
    ws.merge_cells(merge_range(tr, 1, tr, n_fixed))
    c = ws.cell(tr, 1, "TOTAL STUDENTS PER DAY")
    style_cell(
        c, bg=DARK_GREY, fg="FFFFFF", bold=True, h_align="left", border_colour="FFFFFF"
    )
    for di, d in enumerate(ordered_dates):
        n_att = (mat[d] == "✓").sum()
        c = ws.cell(tr, n_fixed + di + 1, n_att)
        style_cell(c, bg=DARK_NAVY, fg="FFFFFF", bold=True, border_colour="FFFFFF")
    ws.row_dimensions[tr].height = 22

    # Column widths and freeze pane
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 13
    for di in range(n_days):
        ws.column_dimensions[get_column_letter(n_fixed + di + 1)].width = 10
    ws.freeze_panes = ws.cell(data_start, n_fixed + 1)

    # ---- Sheet 2: Daily Summary ----------------------------
    ws2 = wb.create_sheet("Daily Summary")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D1")
    c = ws2.cell(1, 1, f"{school} -- Daily Student Count")
    style_cell(c, bg=NAVY, fg="FFFFFF", bold=True, size=13, border_colour="FFFFFF")
    ws2.row_dimensions[1].height = 28

    for ci, h in enumerate(["Week", "Date", "Day of Week", "Students Attending"], 1):
        c = ws2.cell(2, ci, h)
        style_cell(
            c, bg=DARK_NAVY, fg="FFFFFF", bold=True, size=11, border_colour="FFFFFF"
        )
    ws2.row_dimensions[2].height = 22

    for di, d in enumerate(ordered_dates):
        r = 3 + di
        w = week_of[di] - 1
        n = (mat[d] == "✓").sum()
        ws2.cell(r, 1, week_labels[w])
        ws2.cell(r, 2, short_labels[di])
        ws2.cell(r, 3, d.strftime("%A"))
        ws2.cell(r, 4, n)
        for ci in range(1, 4):
            style_cell(ws2.cell(r, ci), bg=WK_BG[w], h_align="left")
        bg_c, bdr_c = COUNT_GREEN if n >= 5 else COUNT_AMBER if n >= 2 else COUNT_RED
        style_cell(ws2.cell(r, 4), bg=bg_c, fg="FFFFFF", bold=True, border_colour=bdr_c)
        ws2.row_dimensions[r].height = 20

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 22

    note = 3 + n_days + 1
    ws2.merge_cells(f"A{note}:D{note}")
    c = ws2.cell(
        note,
        1,
        "Colour key:  Green = 5+ students   Amber = 2-4 students   Red = 0-1 students",
    )
    c.font = Font(name="Arial", size=9, color="555555", italic=True)
    c.alignment = Alignment(horizontal="left")

    return wb


# ============================================================
# MAIN
# ============================================================


def main():
    print()
    print("=" * 55)
    print("  Placement Schedule Generator")
    print("=" * 55)
    print()

    # -- Find the CSV ----------------------------------------
    input_file = "Source.csv"
    if not os.path.exists(input_file):
        print("  'Source.csv' not found in this folder.")
        print("  Type the exact filename of your Google Form export:")
        while True:
            input_file = input("  Filename: ").strip()
            if os.path.exists(input_file):
                break
            print(f"  File '{input_file}' not found. Check spelling and try again.")
    else:
        print(f"  Found '{input_file}' automatically.")

    print()

    # -- Load and auto-detect everything ---------------------
    print("  Reading CSV...")
    raw = load_csv(input_file)
    year = extract_year_from_timestamps(raw)
    print(f"  Placement year detected: {year}")

    ordered_dates, day_lookup = detect_placement_dates(raw, year)

    short_labels = [
        f"{d.strftime('%a')} {d.day} {d.strftime('%b')}" for d in ordered_dates
    ]
    week_of = [i // 5 + 1 for i in range(15)]
    week_labels = []
    for w in range(1, 4):
        idx = [i for i, x in enumerate(week_of) if x == w]
        d1, d2 = ordered_dates[idx[0]], ordered_dates[idx[-1]]
        week_labels.append(
            f"Week {w}  ({d1.strftime('%b')} {d1.day} - {d2.strftime('%b')} {d2.day})"
        )
    range_label = (
        f"{ordered_dates[0].strftime('%b')} {ordered_dates[0].day} - "
        f"{ordered_dates[14].strftime('%b')} {ordered_dates[14].day} "
        f"{ordered_dates[14].year}"
    )

    print(f"  Placement period: {range_label}")
    print(
        f"  Placement days:   {ordered_dates[0].strftime('%d %b')} "
        f"to {ordered_dates[14].strftime('%d %b %Y')} "
        f"({len(ordered_dates)} working days)"
    )
    print(f"  Students:         {len(raw)}")
    print(f"  Schools:          {raw['school'].nunique()}")
    print()

    # -- School selection ------------------------------------
    all_schools = sorted(raw["school"].dropna().unique())
    schools_to_process = ask_school_selection(all_schools)

    # -- Expand to long format -------------------------------
    rows = []
    for _, s in raw.iterrows():
        if pd.isna(s["days_raw"]):
            continue
        for token in str(s["days_raw"]).split(","):
            token = token.strip()
            if token in day_lookup:
                rows.append(
                    {
                        "school": s["school"],
                        "display_name": s["display_name"],
                        "full_name": s["full_name"],
                        "email": s["email"],
                        "student_id": s["student_id"],
                        "date": day_lookup[token],
                    }
                )
    long_df = pd.DataFrame(rows).drop_duplicates()

    # -- Generate Excel files --------------------------------
    output_dir = "school_schedules"
    os.makedirs(output_dir, exist_ok=True)

    n_total = len(schools_to_process)
    print(f"Generating {n_total} Excel file(s)...")
    print()

    for idx, school in enumerate(schools_to_process, 1):
        stu = (
            raw[raw["school"] == school][
                ["display_name", "full_name", "email", "student_id"]
            ]
            .drop_duplicates()
            .sort_values("full_name")
            .reset_index(drop=True)
        )

        school_long = long_df[long_df["school"] == school]
        if not school_long.empty:
            pivot = (
                school_long[["display_name", "date"]]
                .drop_duplicates()
                .assign(present="✓")
                .pivot(index="display_name", columns="date", values="present")
            )
        else:
            pivot = pd.DataFrame(index=[], columns=ordered_dates)

        mat = stu.merge(pivot.reset_index(), on="display_name", how="left")
        for d in ordered_dates:
            if d not in mat.columns:
                mat[d] = ""
        mat[ordered_dates] = mat[ordered_dates].fillna("")

        wb = build_workbook(
            school, mat, ordered_dates, short_labels, week_of, week_labels, range_label
        )

        safe = re.sub(r'[\\/:*?"<>|]', "_", school).strip()
        out_path = os.path.join(output_dir, f"{safe}.xlsx")
        wb.save(out_path)

        if n_total <= 20:
            print(f"  [{idx}/{n_total}]  {school}")
        elif idx % 50 == 0 or idx == n_total:
            print(f"  {idx}/{n_total} done...")

    print()
    print(f"All done!  {n_total} file(s) saved to '{output_dir}/'")
    print()
    input("Press Enter to close.")


if __name__ == "__main__":
    main()
